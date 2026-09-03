#!/usr/bin/env python3
"""
Historical contract notice ingest — 1999 to 2020, from data.gov.au.

The Department of Finance publishes historical CN extracts that carry fields the
live OCDS API does not expose at all:

  * Parent Contract ID, so amendments have real lineage (CN3390353 -> CN3390353-A7)
  * Amendment Date, Amendments Value and Amendment Reason, giving a dated,
    explained value trail
  * SON ID and Panel Arrangement, which is the standing-offer link that makes
    panel and marketplace spend attributable
  * ATM ID, linking a contract back to the approach to market it came from
  * UNSPSC Title, so categories read as words instead of numeric codes
  * Consultancy flags and reasons

Value semantics, verified against CN3360384's 18-row series: Value on an
amendment row is the CUMULATIVE contract value after that amendment, and
Amendments Value is that amendment's delta. 402,745,079 + 162,360 = 402,907,439.

Coverage stops at 2020. This does not help with anything current; it gives a
verified historical baseline, and it demonstrates what Finance is capable of
publishing.

    python historical.py --out data/historical
"""
import argparse, csv, hashlib, io, json, os, re, sys, zipfile
from collections import defaultdict
from datetime import datetime, timezone

import requests

try:
    import openpyxl
except ImportError:
    sys.exit("historical.py needs openpyxl: pip install openpyxl")

UA = os.environ.get("PROCLENS_UA", "LegalTender/1.0 (procurement transparency research)")
TIMEOUT = 300
BASE = ("https://data.gov.au/data/dataset/5c7fa69b-b0e9-4553-b8df-2a022dd2e982"
        "/resource")

# Financial-year extracts. The point-in-time "CNs published up to <date>" files
# overlap these almost entirely and exist for diffing rather than coverage, so
# they are deliberately not ingested here.
SOURCES = [
    ("2019-20", f"{BASE}/06439664-bbcf-4118-a604-164006bffcaa/download/2019-20-australian-government-contract-data.xlsx"),
    ("2018-19", f"{BASE}/ce6b8183-e881-49d4-a367-0851bbb025bf/download/cn-parent-and-amendments-2018-19-30082019.xlsx"),
    ("2017-18", f"{BASE}/bc2097b7-8116-4e9d-9953-98813635892a/download/17-18-fy-dataset.csv"),
    ("2016-17", f"{BASE}/21212500-169f-4745-86b3-6ac1c1174151/download/2016-2017-australian-government-contract-data.csv"),
    ("2014-15", f"{BASE}/561a549b-5a65-450e-86cf-81d392d8fef3/download/20142015fy.csv"),
]

UNSPSC_URL = (f"{BASE}/bae9cb73-1500-4d45-a862-eac2706cfbd4/download/"
              "austender-customised-unspsc-codeset-1-july-2023-1.xlsx")

# Headings drift between years, so match on intent.
COLMAP = [
    ("agency",     r"^agency name$"),
    ("parent",     r"parent contract id"),
    ("cid",        r"^contract id$|^cn id$"),
    ("ntype",      r"contract notice type"),
    ("pub",        r"^publish date$"),
    ("start",      r"^start date$"),
    ("end",        r"^end date$"),
    ("value",      r"^(contract )?value$"),
    ("adate",      r"^amendment date$"),
    ("avalue",     r"amendments? value"),
    ("areason",    r"amendment reason"),
    ("title",      r"^description$"),
    ("agency_ref", r"agency ref"),
    ("unspsc",     r"^unspsc( code)?$"),
    ("cat",        r"unspsc title"),
    ("method",     r"procurement method"),
    ("atm",        r"^atm id$"),
    ("son",        r"^son id$"),
    ("panel",      r"panel arrangement"),
    ("conf",       r"confidentiality contract flag"),
    ("conf_reason", r"confidentiality contract reason"),
    ("conf_out",   r"confidentiality outputs flag"),
    ("consult",    r"consultancy flag"),
    ("consult_reason", r"consultancy reason"),
    ("supplier",   r"supplier name"),
    ("abn",        r"supplier abn$"),
    ("state",      r"supplier state"),
    ("country",    r"supplier country"),
]

FIELDS = ["cn", "agency", "supplier", "abn", "title", "cat", "unspsc", "method",
          "son", "atm", "panel", "conf", "conf_reason", "consult", "consult_reason",
          "state", "country", "agency_ref", "pub", "start", "end",
          "value", "value_first", "amendments", "years"]


def fetch(url):
    r = requests.get(url, headers={"User-Agent": UA}, timeout=TIMEOUT)
    r.raise_for_status()
    return r.content


# Extracts spell "no parent" inconsistently: blank in the 2018-19 workbook, the
# literal string NULL in the 2017-18 CSV. Treating NULL as a real id keyed every
# parent row in that file under one record and collapsed 66,000 contracts into
# a single row.
NULLISH = {"", "null", "none", "n/a", "na", "-", "nil"}


def clean_id(v):
    s = str(v or "").strip()
    return "" if s.lower() in NULLISH else s


def as_money(v):
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    if not v:
        return None
    s = re.sub(r"[^0-9.\-]", "", str(v))
    try:
        return round(float(s), 2)
    except ValueError:
        return None


def as_date(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    if not v:
        return None
    s = str(v).strip()
    for f in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y",
              "%d %b %Y", "%d %B %Y", "%d-%b-%y"):
        try:
            return datetime.strptime(s[:19], f).date().isoformat()
        except ValueError:
            pass
    return s[:10] or None


def map_cols(header):
    idx = {}
    for i, c in enumerate(header):
        name = re.sub(r"\s+", " ", str(c or "")).strip()
        if not name:
            continue
        for key, pat in COLMAP:
            if key not in idx and re.search(pat, name, re.I):
                idx[key] = i
                break
    return idx


def rows_from(blob, url):
    """Yield (header, rowiter) from xlsx, csv or a zip containing a csv."""
    if url.lower().endswith(".xlsx"):
        wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        for r in it:
            if r and sum(1 for c in r if c) > 4:
                return list(r), it
        return None, iter(())
    data = blob
    if url.lower().endswith(".zip"):
        zf = zipfile.ZipFile(io.BytesIO(blob))
        name = next((n for n in zf.namelist() if n.lower().endswith(".csv")), None)
        if not name:
            return None, iter(())
        data = zf.read(name)
    text = data.decode("utf-8-sig", errors="replace")
    rd = csv.reader(io.StringIO(text))
    for r in rd:
        if r and sum(1 for c in r if c) > 4:
            return r, rd
    return None, iter(())


def parse(blob, url, label):
    header, it = rows_from(blob, url)
    if not header:
        return []
    ix = map_cols(header)
    if "cid" not in ix:
        print(f"  ! {label}: no Contract ID column", file=sys.stderr)
        return []
    out = []
    for r in it:
        if not r or not any(x not in (None, "") for x in r):
            continue
        def c(k):
            i = ix.get(k)
            return r[i] if i is not None and i < len(r) else None
        cid = clean_id(c("cid"))
        if not cid or cid.lower() in ("contract id", "cn id"):
            continue
        parent = clean_id(c("parent"))
        # An amendment id is the parent id with an -A<n> suffix; fall back to
        # stripping it when Parent Contract ID is absent or NULL.
        if not parent:
            parent = re.sub(r"-A\d+$", "", cid, flags=re.I)
        ntype = str(c("ntype") or "").strip().lower()
        out.append({
            "cid": cid, "parent": parent,
            "agency": str(c("agency") or "").strip(),
            "supplier": str(c("supplier") or "").strip(),
            "abn": re.sub(r"\D", "", str(c("abn") or "")),
            "title": str(c("title") or "").strip(),
            "cat": str(c("cat") or "").strip(),
            "unspsc": str(c("unspsc") or "").strip(),
            "method": str(c("method") or "").strip(),
            "son": str(c("son") or "").strip(),
            "atm": str(c("atm") or "").strip(),
            "panel": str(c("panel") or "").strip(),
            "conf": str(c("conf") or "").strip(),
            "conf_reason": str(c("conf_reason") or "").strip()[:160],
            "consult": str(c("consult") or "").strip(),
            "consult_reason": str(c("consult_reason") or "").strip()[:160],
            "state": str(c("state") or "").strip(),
            "country": str(c("country") or "").strip(),
            "agency_ref": str(c("agency_ref") or "").strip(),
            "pub": as_date(c("pub")), "start": as_date(c("start")),
            "end": as_date(c("end")),
            "value": as_money(c("value")),
            "ntype": ntype,
            "adate": as_date(c("adate")),
            "avalue": as_money(c("avalue")),
            "areason": str(c("areason") or "").strip()[:140],
            "year": label,
        })
    return out


# A contract id looks like CN1234567, optionally with an -A<n> amendment suffix.
# Anything else is a parsing failure, and letting it through is how thousands of
# unrelated contracts end up merged under one bogus key.
CN_RE = re.compile(r"^(CN|SON)[0-9][0-9A-Za-z._/-]*$", re.I)


def merge(store, rows, label):
    """Collapse to one record per PARENT contract, never duplicating.

    A file holds a parent row plus one row per amendment, and the yearly files
    overlap each other, so the same contract and even the same amendment arrive
    repeatedly. Contracts are keyed on the parent id and amendments on their own
    contract id, so re-ingesting anything is a no-op.
    """
    added = merged = amend_new = rejected = 0
    for r in rows:
        key = r["parent"]
        if not CN_RE.match(key or ""):
            rejected += 1
            continue
        rec = store.get(key)
        if rec is None:
            rec = {"cn": key, "amendments": {}, "years": []}
            store[key] = rec
            added += 1
        else:
            merged += 1
        if label not in rec["years"]:
            rec["years"].append(label)

        # Prefer the file's own Contract Notice Type where it exists; fall back
        # to comparing ids.
        nt = r.get("ntype") or ""
        is_amendment = (nt == "amendment") if nt in ("parent", "amendment") \
                       else (r["cid"] != key)
        if is_amendment:
            # Keyed by the amendment's own id: the same amendment appearing in
            # two yearly files is stored once.
            if r["cid"] not in rec["amendments"]:
                rec["amendments"][r["cid"]] = {
                    "id": r["cid"], "date": r["adate"],
                    "delta": r["avalue"], "total": r["value"],
                    "reason": r["areason"],
                }
                amend_new += 1
        else:
            # The parent row carries the original value.
            if r["value"] is not None:
                rec["value_first"] = r["value"]

        # Descriptive fields: fill gaps, prefer a populated later value.
        for f in ("agency", "supplier", "abn", "title", "cat", "unspsc", "method",
                  "son", "atm", "panel", "conf", "conf_reason", "consult",
                  "consult_reason", "state", "country", "agency_ref",
                  "pub", "start", "end"):
            if r.get(f) not in (None, "") and not rec.get(f):
                rec[f] = r[f]
    if rejected:
        print(f"    ! {rejected:,} rows rejected: parent id did not look like a "
              f"contract id", file=sys.stderr)
    return added, merged, amend_new


def amend_seq(a):
    """Order by the -A<n> suffix; amendment dates are not always monotonic."""
    m = re.search(r"-A(\d+)$", str(a.get("id") or ""), re.I)
    return int(m.group(1)) if m else 0


def finalise(store):
    """Resolve current and original value from the amendment trail."""
    out = []
    for rec in store.values():
        am = sorted(rec.pop("amendments", {}).values(),
                    key=lambda a: (amend_seq(a), a.get("date") or ""))
        if am:
            # The parent row's Value is the contract's CURRENT total at the time
            # of that extract, not its original award value: parent CN3390353
            # and its amendment -A7 both carry $3,964,263,051. So the original
            # is recoverable only from the earliest amendment that publishes
            # both a cumulative and a delta.
            first = next((a for a in am if a.get("total") is not None
                          and a.get("delta") is not None), None)
            rec["value_first"] = round(first["total"] - first["delta"], 2) \
                if first is not None else None

            # Cumulative totals are deliberately NOT reconstructed where the
            # source omits them. The 2017-18 extract publishes only deltas, and
            # accumulating them onto an assumed base produced $898,199,987 for
            # CN3360384 against a published $780,644,545. A figure that
            # contradicts the source is worse than an absent one.
            totals = [a["total"] for a in am if a.get("total") is not None]
            if totals:
                rec["value"] = totals[-1]
            for a in am:
                a["derived"] = 0
        rec["amendments"] = [[a["id"], a["date"], a["delta"], a["total"],
                              a["reason"], a.get("derived", 0)] for a in am]
        if rec.get("value") is None:
            rec["value"] = rec.get("value_first")
        rec["years"] = sorted(rec.get("years") or [])
        out.append(rec)
    out.sort(key=lambda r: r.get("value") or 0, reverse=True)
    return out


DICT_FIELDS = ("agency", "supplier", "cat", "method", "panel", "conf",
               "consult", "state", "country")


def encode(rows):
    dicts, lookup = {}, {}
    for f in DICT_FIELDS:
        vals, seen = [], {}
        for r in rows:
            v = r.get(f)
            if v not in seen:
                seen[v] = len(vals)
                vals.append(v)
        dicts[f], lookup[f] = vals, seen
    return dicts, [[lookup[f][r.get(f)] if f in lookup else r.get(f) for f in FIELDS]
                   for r in rows]


def read_json(path, default=None):
    try:
        with open(path) as fh:
            return json.load(fh)
    except (OSError, json.JSONDecodeError):
        return default if default is not None else {}


def write_json_if_changed(path, obj):
    blob = json.dumps(obj, separators=(",", ":"), default=str)
    stable = {k: v for k, v in obj.items() if k != "generated"}
    try:
        with open(path) as fh:
            old = json.load(fh)
        if {k: v for k, v in old.items() if k != "generated"} == stable:
            return False
    except (OSError, json.JSONDecodeError, AttributeError):
        pass
    tmp = path + ".tmp"
    with open(tmp, "w") as fh:
        fh.write(blob)
    os.replace(tmp, path)
    return True


def shard_year(r):
    p = r.get("pub") or r.get("start") or ""
    return p[:4] if len(p) >= 4 and p[:4].isdigit() else "unknown"


def load_store(outdir):
    """Read every year shard back into one CN-keyed store."""
    store = {}
    if not os.path.isdir(outdir):
        return store
    for name in sorted(os.listdir(outdir)):
        if name.startswith("contracts-") and name.endswith(".json"):
            store.update(_load_one(os.path.join(outdir, name)))
    # Migration path: an earlier build wrote a single contracts.json. Read it so
    # a change of layout never silently orphans an already-ingested store —
    # sources are hash-skipped on a re-run, so nothing would rebuild it.
    legacy = os.path.join(outdir, "contracts.json")
    if os.path.exists(legacy):
        store.update(_load_one(legacy))
    return store


def _load_one(path):
    b = read_json(path, {})
    fields = b.get("fields", FIELDS)
    dicts = b.get("dict") or {}
    store = {}
    for row in b.get("rows", []):
        r = dict(zip(fields, row))
        for f, vals in dicts.items():
            i = r.get(f)
            if isinstance(i, int) and 0 <= i < len(vals):
                r[f] = vals[i]
        # Re-key amendments so a re-run merges rather than appends.
        r["amendments"] = {a[0]: {"id": a[0], "date": a[1], "delta": a[2],
                                  "total": a[3], "reason": a[4]}
                           for a in (r.get("amendments") or [])}
        if r.get("cn"):
            store[r["cn"]] = r
    return store


def unspsc_map():
    try:
        blob = fetch(UNSPSC_URL)
    except Exception as e:
        print(f"! UNSPSC codeset: {e}", file=sys.stderr)
        return {}
    wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    out = {}
    for r in ws.iter_rows(values_only=True):
        if not r or len(r) < 2:
            continue
        code = re.sub(r"\D", "", str(r[0] or ""))
        title = str(r[1] or "").strip()
        if code and title and not title.lower().startswith("unspsc"):
            out[code] = title
    return out


def main():
    ap = argparse.ArgumentParser(description="Ingest historical CN extracts.")
    ap.add_argument("--out", default="data/historical")
    ap.add_argument("--only", help="restrict to one year label, e.g. 2018-19")
    ap.add_argument("--force", action="store_true", help="re-parse ingested files")
    ap.add_argument("--skip-unspsc", action="store_true")
    args = ap.parse_args()
    os.makedirs(args.out, exist_ok=True)

    index_path = os.path.join(args.out, "index.json")
    store = load_store(args.out)
    index = read_json(index_path, {})
    done = {s["sha"]: s for s in index.get("sources", [])}
    print(f"Store holds {len(store):,} contracts from "
          f"{len(index.get('sources', []))} files.", file=sys.stderr)

    if not args.skip_unspsc:
        m = unspsc_map()
        if m:
            write_json_if_changed(os.path.join(args.out, "unspsc.json"),
                                  {"generated": datetime.now(timezone.utc)
                                   .isoformat(timespec="seconds"), "codes": m})
            print(f"  UNSPSC codeset: {len(m):,} codes", file=sys.stderr)

    sources = list(index.get("sources", []))
    for label, url in SOURCES:
        if args.only and label != args.only:
            continue
        try:
            blob = fetch(url)
        except Exception as e:
            print(f"! {label}: {e}", file=sys.stderr)
            continue
        sha = hashlib.sha256(blob).hexdigest()[:16]
        if sha in done and not args.force:
            print(f"  {label}: already ingested, skipped", file=sys.stderr)
            continue
        rows = parse(blob, url, label)
        if not rows:
            print(f"  ! {label}: no rows parsed", file=sys.stderr)
            continue
        added, merged_n, amend_new = merge(store, rows, label)
        rec = {"label": label, "url": url, "sha": sha, "rows": len(rows),
               "new_contracts": added, "merged_into_existing": merged_n,
               "new_amendments": amend_new,
               "captured": datetime.now(timezone.utc).date().isoformat()}
        sources = [s for s in sources if s.get("sha") != sha and s.get("label") != label]
        sources.append(rec)
        done[sha] = rec
        print(f"  {label}: {len(rows):,} rows -> +{added:,} contracts, "
              f"{merged_n:,} merged into existing, +{amend_new:,} amendments",
              file=sys.stderr)

    rows_out = finalise(store)
    # One 71 MB file is too much to hand a browser. Sharded by publication year
    # so a date range only pulls the years it needs, and each shard carries the
    # count and byte size the front end uses to bust its cache.
    by_year = defaultdict(list)
    for r in rows_out:
        by_year[shard_year(r)].append(r)
    shards = []
    for year in sorted(by_year):
        rows = by_year[year]
        dicts, encoded = encode(rows)
        name = f"contracts-{year}.json"
        path = os.path.join(args.out, name)
        write_json_if_changed(path, {"year": year, "fields": FIELDS,
                                     "dict": dicts, "rows": encoded})
        shards.append({"year": year, "file": name, "count": len(rows),
                       "bytes": os.path.getsize(path)})
    stale = os.path.join(args.out, "contracts.json")
    if shards and os.path.exists(stale):
        os.remove(stale)

    with_am = sum(1 for r in rows_out if r.get("amendments"))
    with_son = sum(1 for r in rows_out if r.get("son"))
    with_atm = sum(1 for r in rows_out if r.get("atm"))
    sources.sort(key=lambda s: str(s.get("label")), reverse=True)
    write_json_if_changed(index_path, {
        "generated": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "note": "Historical Australian Government contract notices, 1999-2020, "
                "published by the Department of Finance on data.gov.au. Carries "
                "amendment lineage, SON and ATM ids, panel flags and UNSPSC "
                "titles, none of which the live OCDS API exposes. Coverage ends "
                "at 2020.",
        "sources": sources,
        "fields": FIELDS,
        "shards": shards,
        "totals": {"contracts": len(rows_out), "with_amendments": with_am,
                   "with_son_id": with_son, "with_atm_id": with_atm},
    })
    print(f"Store now {len(rows_out):,} contracts; {with_am:,} with amendment "
          f"history, {with_son:,} with a SON id, {with_atm:,} with an ATM id.",
          file=sys.stderr)


if __name__ == "__main__":
    main()
