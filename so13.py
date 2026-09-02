#!/usr/bin/env python3
"""
Senate Order 13 ingest — contracts from entities that do not report to AusTender.

Corporate Commonwealth entities are not bound by the Commonwealth Procurement
Rules' reporting requirements, so their contracts never appear in AusTender. The
NDIA is the significant example: it spends billions and publishes nothing to the
OCDS feed. Under Senate Continuing Order 13 those entities must instead publish,
twice yearly, every contract of $100,000 or more (GST inclusive) on their own
website, as a spreadsheet.

This ingests those spreadsheets.

Two things make this data BETTER than AusTender for value questions:
  * it carries "Original Contract Value" alongside the current consideration,
    which is the amendment history the OCDS API refuses to give; and
  * it states explicitly whether a contract carries confidentiality provisions,
    and why.

    python so13.py --out data/so13
"""
import argparse, io, json, os, re, sys
from datetime import datetime, timezone

import requests

try:
    import openpyxl
except ImportError:
    sys.exit("so13.py needs openpyxl: pip install openpyxl")

UA = os.environ.get("PROCLENS_UA", "LegalTender/1.0 (procurement transparency research)")
TIMEOUT = 90

# Entities that publish Senate Order 13 listings instead of reporting to
# AusTender. Add to this list as more are identified; the parser is generic.
SOURCES = [
    {
        "entity": "National Disability Insurance Agency",
        "short": "ndia",
        "page": "https://www.ndis.gov.au/policies-rules-and-legal/legal/senate-order-13-entity-contracts",
        "base": "https://www.ndis.gov.au",
    },
]

# Column headings vary between entities and between periods, so match on intent
# rather than on an exact string.
COLMAP = [
    ("supplier",      r"contractor|supplier|vendor|company"),
    ("abn",           r"\babn\b"),
    ("title",         r"subject matter|description|subject"),
    ("start",         r"commencement|start date"),
    ("end",           r"anticipated end|end date|expiry"),
    ("confidential",  r"provisions requiring the parties to maintain"),
    ("conf_reason",   r"^reason ?\(s\)$"),
    ("conf_other",    r"other requirements of confidentiality"),
    ("value",         r"consideration"),
    ("cn",            r"contract id|contract number"),
    ("ctype",         r"contract type"),
    ("method",        r"procurement method"),
    ("approached",    r"suppliers approached"),
    ("value_first",   r"original contract value"),
    ("variations",    r"number of variation"),
]

FIELDS = ["entity", "supplier", "abn", "title", "value", "value_first", "variations",
          "start", "end", "method", "ctype", "approached", "confidential",
          "conf_reason", "cn", "period", "source_url"]


def get(url, binary=False):
    r = requests.get(url, headers={"User-Agent": UA}, timeout=TIMEOUT)
    r.raise_for_status()
    return r.content if binary else r.text


def discover(src):
    """Find the listing spreadsheets linked from an entity's SO13 page."""
    html = get(src["page"])
    out, seen = [], set()
    # Links are media ids; the visible text carries the reporting period.
    for m in re.finditer(r'href="(/media/(\d+)/download[^"]*)"([^>]*)>(.{0,400}?)</a>',
                         html, re.S | re.I):
        href, mid, _, label = m.groups()
        if mid in seen:
            continue
        seen.add(mid)
        text = re.sub(r"<[^>]+>", " ", label)
        text = re.sub(r"\s+", " ", text).strip()
        if not re.search(r"xlsx|excel|csv", text, re.I) and not re.search(r"20\d\d", text):
            continue
        out.append({"url": src["base"] + href, "label": text[:120],
                    "period": period_of(text), "xlsx": bool(re.search(r"xlsx|excel", text, re.I))})
    return out


def period_of(text):
    m = re.search(r"(20\d\d)\s*[-–—]\s*(?:20)?(\d\d)", text)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    m = re.search(r"(20\d\d)", text)
    return m.group(1) if m else "unknown"


def header_row(rows):
    for i, r in enumerate(rows[:25]):
        if r and sum(1 for c in r if c) > 4:
            return i
    return None


def map_columns(cols):
    idx = {}
    for i, c in enumerate(cols):
        name = re.sub(r"\s+", " ", str(c or "")).strip()
        if not name:
            continue
        for key, pat in COLMAP:
            if key in idx:
                continue
            if re.search(pat, name, re.I):
                idx[key] = i
                break
    return idx


def as_money(v):
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    if not v:
        return None
    s = re.sub(r"[^0-9.\-]", "", str(v))
    try:
        return round(float(s), 2)
    except ValueError:
        return None


def as_date(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    if not v:
        return None
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d %b %Y", "%d %B %Y"):
        try:
            return datetime.strptime(s[:20], fmt).date().isoformat()
        except ValueError:
            pass
    return s[:10] or None


def parse(blob, entity, period, url):
    wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    sheet = next((n for n in wb.sheetnames if re.search(r"contract|listing", n, re.I)),
                 wb.sheetnames[-1])
    rows = [r for r in wb[sheet].iter_rows(values_only=True)]
    h = header_row(rows)
    if h is None:
        return []
    idx = map_columns(rows[h])
    if "supplier" not in idx:
        return []
    out = []
    for r in rows[h + 1:]:
        if not r or not any(r):
            continue
        def cell(k):
            i = idx.get(k)
            return r[i] if i is not None and i < len(r) else None
        supplier = str(cell("supplier") or "").strip()
        if not supplier or re.match(r"^(total|contractor)$", supplier, re.I):
            continue
        out.append({
            "entity": entity,
            "supplier": supplier,
            "abn": re.sub(r"\D", "", str(cell("abn") or "")) or "",
            "title": str(cell("title") or "").strip(),
            "value": as_money(cell("value")),
            "value_first": as_money(cell("value_first")),
            "variations": cell("variations"),
            "start": as_date(cell("start")),
            "end": as_date(cell("end")),
            "method": str(cell("method") or "").strip(),
            "ctype": str(cell("ctype") or "").strip(),
            "approached": cell("approached"),
            "confidential": str(cell("confidential") or "").strip(),
            "conf_reason": str(cell("conf_reason") or "").strip(),
            "cn": str(cell("cn") or "").strip(),
            "period": period,
            "source_url": url,
        })
    return out


def main():
    ap = argparse.ArgumentParser(description="Ingest Senate Order 13 contract listings.")
    ap.add_argument("--out", default="data/so13")
    ap.add_argument("--only", help="restrict to one entity short name, e.g. ndia")
    args = ap.parse_args()
    os.makedirs(args.out, exist_ok=True)

    manifest, total = [], 0
    for src in SOURCES:
        if args.only and src["short"] != args.only:
            continue
        try:
            files = discover(src)
        except Exception as e:
            print(f"! {src['short']}: cannot read page ({e})", file=sys.stderr)
            continue
        print(f"{src['short']}: {len(files)} listing files found", file=sys.stderr)
        for f in files:
            if not f["xlsx"]:
                # PDF listings exist for older periods; they need a different
                # extractor and are skipped rather than guessed at.
                print(f"  skip (not a spreadsheet): {f['label']}", file=sys.stderr)
                continue
            try:
                rows = parse(get(f["url"], binary=True), src["entity"], f["period"], f["url"])
            except Exception as e:
                print(f"  ! {f['label']}: {e}", file=sys.stderr)
                continue
            if not rows:
                print(f"  ! {f['label']}: no rows parsed", file=sys.stderr)
                continue
            name = f"{src['short']}-{f['period']}.json"
            path = os.path.join(args.out, name)
            payload = {"entity": src["entity"], "period": f["period"],
                       "source_url": f["url"], "label": f["label"],
                       "fields": FIELDS,
                       "rows": [[r.get(k) for k in FIELDS] for r in rows]}
            blob = json.dumps(payload, separators=(",", ":"), default=str)
            old = None
            try:
                with open(path) as fh:
                    old = fh.read()
            except OSError:
                pass
            if old != blob:
                with open(path, "w") as fh:
                    fh.write(blob)
            val = sum(r["value"] for r in rows if r.get("value"))
            manifest.append({"entity": src["entity"], "short": src["short"],
                             "period": f["period"], "file": name, "count": len(rows),
                             "total": round(val, 2), "source_url": f["url"]})
            total += len(rows)
            print(f"  {f['period']}: {len(rows)} contracts, ${val:,.0f}", file=sys.stderr)

    manifest.sort(key=lambda m: (m["short"], m["period"]), reverse=True)
    with open(os.path.join(args.out, "index.json"), "w") as fh:
        json.dump({
            "generated": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "note": "Senate Order 13 listings from entities that do not report to "
                    "AusTender. Values are GST inclusive and cover contracts of "
                    "$100,000 or more only, so smaller spend is absent by design.",
            "fields": FIELDS,
            "listings": manifest,
        }, fh, separators=(",", ":"))
    print(f"Total {total} contracts across {len(manifest)} listings.", file=sys.stderr)


if __name__ == "__main__":
    main()
