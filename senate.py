#!/usr/bin/env python3
"""
Senate Order snapshot capture — preserving what AusTender deletes.

AusTender publishes a point-in-time report, twice yearly, of every procurement
contract of $100,000 or more that was active during the period. It carries two
things the OCDS API does not expose at all:

  * whether the contract has confidentiality provisions, and the stated reason; and
  * a CN ID, which joins these rows straight onto the OCDS archive.

The catch is in AusTender's own words: the archive holds "a maximum three
periods". Probing SnapshotId 8-24 confirms it — everything outside the three
linked ids returns an identical error page, so old snapshots are deleted rather
than merely unlinked. The last five years are already gone. The only way to hold
five years is to capture each period before it rolls off, which is what this does.

Because each snapshot is a point in time, the same contract recurring across
periods at a different value is real, dated evidence of a variation — the
amendment history the OCDS API has no endpoint for.

    python senate.py --out data/senate
"""
import argparse, hashlib, io, json, os, re, sys
from datetime import datetime, timezone

import requests

try:
    import openpyxl
except ImportError:
    sys.exit("senate.py needs openpyxl: pip install openpyxl")

BASE = "https://www.tenders.gov.au"
LIST = BASE + "/senateorder/list"
ARCHIVED = BASE + "/senateorder/archivedlist"

# tenders.gov.au rejects plain requests with 403. It is a header check, not a
# real block: a browser User-Agent plus a Referer is sufficient.
HEADERS = {
    "User-Agent": ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                   "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0 Safari/537.36"),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}
TIMEOUT = 120

COLMAP = [
    ("agency",        r"agency name|entity name"),
    ("cn",            r"^cn id$|contract notice"),
    ("supplier",      r"supplier name"),
    ("abn",           r"supplier abn|^abn$"),
    ("title",         r"^description$"),
    ("cat",           r"^category$"),
    ("agency_ref",    r"agency ref"),
    ("conf_contract", r"confidentiality - contract"),
    ("conf_reason",   r"confidentiality reason\(s\) - contract"),
    ("conf_outputs",  r"confidentiality - outputs"),
    ("conf_out_reason", r"confidentiality reason\(s\) - outputs"),
    ("pub",           r"publish date"),
    ("start",         r"start date"),
    ("end",           r"end date"),
    ("value",         r"^value"),
]

# One record per contract; observations carry the per-period history.
FIELDS = ["cn", "agency", "supplier", "abn", "title", "cat", "agency_ref",
          "conf_contract", "conf_reason", "conf_outputs", "conf_out_reason",
          "pub", "start", "end", "value", "value_first", "periods",
          "observations", "first_seen", "last_seen"]


def fetch(url, referer=LIST):
    h = dict(HEADERS)
    h["Referer"] = referer
    r = requests.get(url, headers=h, timeout=TIMEOUT)
    r.raise_for_status()
    return r


def discover():
    """Current All Agencies report plus every archived snapshot still served."""
    out = [{"id": "current", "url": BASE + "/SenateOrder/Download", "label": "current"}]
    try:
        html = fetch(LIST).text
        m = re.search(r"Agency Name.*?All Agencies.*?<td[^>]*>([^<]{4,40})</td>", html, re.S | re.I)
        if m:
            out[0]["label"] = re.sub(r"\s+", " ", m.group(1)).strip()
    except Exception as e:
        print(f"! cannot read current list ({e})", file=sys.stderr)
    try:
        html = fetch(ARCHIVED).text
        for m in re.finditer(r'href="([^"]*SenateOrder/Download\?SnapshotId=(\d+))"[^>]*>(.{0,120}?)</a>',
                             html, re.S | re.I):
            href, sid, label = m.groups()
            label = re.sub(r"\s+", " ", re.sub("<[^>]+>", " ", label)).strip()
            url = href if href.startswith("http") else BASE + href
            out.append({"id": sid, "url": url, "label": label})
    except Exception as e:
        print(f"! cannot read archived list ({e})", file=sys.stderr)
    return out


def as_money(v):
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    if not v:
        return None
    s = re.sub(r"[^0-9.\-]", "", str(v))
    try:
        return round(float(s), 2)
    except ValueError:
        return None


def as_date(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    if not v:
        return None
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d %b %Y", "%d %B %Y"):
        try:
            return datetime.strptime(s[:20], fmt).date().isoformat()
        except ValueError:
            pass
    return s[:10] or None


def parse(blob):
    wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    h = next((i for i, r in enumerate(rows[:40]) if r and sum(1 for c in r if c) > 4), None)
    if h is None:
        return None, []
    idx = {}
    for i, c in enumerate(rows[h]):
        name = re.sub(r"\s+", " ", str(c or "")).strip()
        for key, pat in COLMAP:
            if key not in idx and name and re.search(pat, name, re.I):
                idx[key] = i
                break
    if "cn" not in idx:
        return None, []
    out = []
    for r in rows[h + 1:]:
        if not r or not any(r):
            continue
        def cell(k):
            i = idx.get(k)
            return r[i] if i is not None and i < len(r) else None
        cn = str(cell("cn") or "").strip()
        if not cn or cn.lower() == "cn id":
            continue
        out.append({
            "cn": cn,
            "agency": str(cell("agency") or "").strip(),
            "supplier": str(cell("supplier") or "").strip(),
            "abn": re.sub(r"\D", "", str(cell("abn") or "")),
            "title": str(cell("title") or "").strip(),
            "cat": str(cell("cat") or "").strip(),
            "agency_ref": str(cell("agency_ref") or "").strip(),
            "conf_contract": str(cell("conf_contract") or "").strip(),
            "conf_reason": str(cell("conf_reason") or "").strip(),
            "conf_outputs": str(cell("conf_outputs") or "").strip(),
            "conf_out_reason": str(cell("conf_out_reason") or "").strip(),
            "pub": as_date(cell("pub")),
            "start": as_date(cell("start")),
            "end": as_date(cell("end")),
            "value": as_money(cell("value")),
        })
    return idx, out


def load_store(path):
    b = read_json(path, {})
    fields = b.get("fields", FIELDS)
    store = {}
    for row in b.get("rows", []):
        r = dict(zip(fields, row))
        if r.get("cn"):
            store[r["cn"]] = r
    return store


def merge(store, rows, period, today):
    """Fold one snapshot in, without duplicating.

    A contract active across several periods appears in every one of those
    snapshots. It is stored once, keyed by CN ID; a period is recorded against
    it only once, and a new observation is appended only when the value actually
    moved. Re-running over a snapshot already ingested therefore changes nothing.
    """
    added = updated = repeats = 0
    for r in rows:
        cn = r["cn"]
        old = store.get(cn)
        if old is None:
            r["value_first"] = r.get("value")
            r["periods"] = [period]
            r["observations"] = [[period, r.get("value")]]
            r["first_seen"] = today
            r["last_seen"] = today
            store[cn] = r
            added += 1
            continue

        periods = list(old.get("periods") or [])
        obs = [list(o) for o in (old.get("observations") or [])]
        seen_before = period in periods
        if not seen_before:
            periods.append(period)

        # Append an observation only on a real value move, so re-ingesting the
        # same snapshot cannot inflate the history.
        last_val = obs[-1][1] if obs else None
        moved = r.get("value") is not None and r.get("value") != last_val
        if moved and not any(o[0] == period for o in obs):
            obs.append([period, r.get("value")])

        merged = dict(old)
        # Later snapshots supersede earlier detail, but never the first value.
        for k in ("agency", "supplier", "abn", "title", "cat", "agency_ref",
                  "conf_contract", "conf_reason", "conf_outputs",
                  "conf_out_reason", "pub", "start", "end", "value"):
            if r.get(k) not in (None, ""):
                merged[k] = r[k]
        obs.sort(key=lambda o: period_end(o[0]))
        periods.sort(key=period_end)
        merged["value_first"] = obs[0][1] if obs else old.get("value_first")
        merged["periods"] = periods
        merged["observations"] = obs
        merged["first_seen"] = old.get("first_seen") or today
        merged["last_seen"] = today
        store[cn] = merged
        if seen_before:
            repeats += 1
        elif moved:
            updated += 1
    return added, updated, repeats


def read_json(path, default=None):
    try:
        with open(path) as fh:
            return json.load(fh)
    except (OSError, json.JSONDecodeError):
        return default if default is not None else {}


def write_json_if_changed(path, obj):
    """Compare ignoring `generated`, or the timestamp alone would force a commit
    on every run even when not a single contract changed."""
    blob = json.dumps(obj, separators=(",", ":"), default=str)
    stable = {k: v for k, v in obj.items() if k != "generated"}
    try:
        with open(path) as fh:
            old = json.load(fh)
        if {k: v for k, v in old.items() if k != "generated"} == stable:
            return False
    except (OSError, json.JSONDecodeError, AttributeError):
        pass
    tmp = path + ".tmp"
    with open(tmp, "w") as fh:
        fh.write(blob)
    os.replace(tmp, path)
    return True


def main():
    ap = argparse.ArgumentParser(description="Capture AusTender Senate Order snapshots.")
    ap.add_argument("--out", default="data/senate")
    ap.add_argument("--force", action="store_true",
                    help="re-parse snapshots already ingested")
    args = ap.parse_args()
    os.makedirs(args.out, exist_ok=True)

    store_path = os.path.join(args.out, "contracts.json")
    index_path = os.path.join(args.out, "index.json")
    store = load_store(store_path)
    index = read_json(index_path, {})
    done = {s["sha"]: s for s in index.get("snapshots", [])}
    by_id = {s["id"]: s for s in index.get("snapshots", [])}
    today = datetime.now(timezone.utc).date().isoformat()

    print(f"Store holds {len(store)} contracts from "
          f"{len(index.get('snapshots', []))} snapshots.", file=sys.stderr)

    snapshots = list(index.get("snapshots", []))
    # Oldest first, so value_first really is the earliest observed value and the
    # observation trail reads forwards in time.
    found = sorted(discover(), key=lambda x: period_end(period_of(x["label"])))
    for s in found:
        try:
            resp = fetch(s["url"], referer=ARCHIVED)
        except Exception as e:
            print(f"! {s['id']}: {e}", file=sys.stderr)
            continue
        blob = resp.content
        sha = hashlib.sha256(blob).hexdigest()[:16]
        # Content hash is the duplicate guard: the "current" report becomes an
        # archived one under a new id, so the same bytes reappear under a
        # different URL and must not be ingested twice.
        if sha in done and not args.force:
            print(f"  {s['label'] or s['id']}: already ingested, skipped",
                  file=sys.stderr)
            continue
        idx, rows = parse(blob)
        if not rows:
            print(f"  ! {s['label'] or s['id']}: no rows parsed", file=sys.stderr)
            continue
        period = period_of(s["label"]) or s["id"]
        added, updated, repeats = merge(store, rows, period, today)
        rec = {"id": s["id"], "period": period, "label": s["label"],
               "url": s["url"], "sha": sha, "rows": len(rows),
               "captured": today, "entities": len({r["agency"] for r in rows if r["agency"]})}
        snapshots = [x for x in snapshots if x.get("sha") != sha and x.get("id") != s["id"]]
        snapshots.append(rec)
        done[sha] = rec
        print(f"  {period}: {len(rows)} rows, {rec['entities']} entities "
              f"(+{added} new, {updated} value moves, {repeats} already-known)",
              file=sys.stderr)

    rows_out = sorted(store.values(), key=lambda r: r.get("value") or 0, reverse=True)
    write_json_if_changed(store_path,
                          {"generated": datetime.now(timezone.utc).isoformat(timespec="seconds"),
                           "fields": FIELDS,
                           "rows": [[r.get(f) for f in FIELDS] for r in rows_out]})

    conf = sum(1 for r in store.values()
               if str(r.get("conf_contract") or "").upper().startswith("Y"))
    multi = sum(1 for r in store.values() if len(r.get("observations") or []) > 1)
    snapshots.sort(key=lambda s: str(s.get("period")), reverse=True)
    write_json_if_changed(index_path, {
        "generated": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "note": "Point-in-time Senate Order snapshots of Commonwealth procurement "
                "contracts of $100,000 or more. AusTender keeps only three periods "
                "and deletes the rest; these are preserved permanently. Rows join "
                "to the OCDS archive on CN ID.",
        "snapshots": snapshots,
        "totals": {"contracts": len(store),
                   "with_confidentiality": conf,
                   "with_value_history": multi},
    })
    print(f"Store now {len(store)} contracts; {conf} carry confidentiality "
          f"provisions; {multi} have multi-period value history.", file=sys.stderr)


def period_end(period):
    """Sort key: the date a period closes, so CY2024 < FY2024-25 < CY2025."""
    if not period:
        return "9999"
    m = re.match(r"FY(20\d\d)-(\d\d)$", period)
    if m:
        return f"20{m.group(2)}-06-30"
    m = re.match(r"CY(20\d\d)$", period)
    if m:
        return f"{m.group(1)}-12-31"
    return "9999"


def period_of(label):
    if not label:
        return None
    m = re.search(r"(20\d\d)\s*/\s*(20\d\d)\s*financial", label, re.I)
    if m:
        return f"FY{m.group(1)}-{m.group(2)[2:]}"
    m = re.search(r"(20\d\d)\s*calendar", label, re.I)
    if m:
        return f"CY{m.group(1)}"
    return None


if __name__ == "__main__":
    main()
